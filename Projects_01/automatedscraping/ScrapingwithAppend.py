import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import schedule
import time
import os

def scraping():
    url = "https://merolagani.com/latestmarket.aspx"
    res = requests.get(url)
    soup = BeautifulSoup(res.text, 'html.parser')

    table = soup.find('table')
    rows = table.find_all('tr')[1:]

    data = []
    for row in rows:
        cols = row.find_all('td')
        if len(cols) >= 4:
            symbol = cols[0].text.strip()
            ltp = cols[1].text.strip()
            open_price = cols[3].text.strip()
            data.append((symbol, ltp, open_price))
    return pd.DataFrame(data, columns=["Symbol", "LTP", "Open"])

def wr_to_excel(data, filename):
    if os.path.exists(filename):
        book = load_workbook(filename)
        sheet = book['Sheet1']
        startrow = sheet.max_row

        with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists='overlay') as writer:
            data.to_excel(writer, index=False, sheet_name='Sheet1', startrow=startrow, header=False)
    else:
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            data.to_excel(writer, index=False, sheet_name='Sheet1')

def job():
    print("Running scraping job...")
    data = scraping()
    wr_to_excel(data, "data.xlsx")
    print("Job done.")

job()

schedule.every(10).minutes.do(job)

if __name__ == "__main__":
    while True:
        schedule.run_pending()
        time.sleep(1)